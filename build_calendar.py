#!/usr/bin/env python3
"""
build_calendar.py – Generate a styled didactics calendar HTML page
from “Didactics _ UCI Derm 2017-Present (2).xlsx”.

Run:
    python build_calendar.py "Didactics _ UCI Derm 2017-Present (2).xlsx"  didactics_calendar.html

Requires:
    pip install openpyxl
"""
from __future__ import annotations

import argparse, calendar, datetime as dt, html, math, re, sys
from pathlib import Path

import openpyxl


# ────────────────────────── utility helpers ──────────────────────────
def format_time(val) -> str:
    """Return clean '8–9 AM' style strings, coping with Excel quirks."""
    if val is None:
        return ""
    if isinstance(val, str):
        s = re.sub(r"\s*[-–]\s*", "–", val.strip())
        if not re.search(r"\b(?:am|pm)\b", s, flags=re.I):
            s += " AM"
        return s
    if isinstance(val, (dt.datetime, dt.date)):            # Excel mis-parsed “8-9”
        return f"{val.month}–{val.day} AM"
    if isinstance(val, (float, int)):                      # Serial time
        minutes = round(val * 24 * 60)
        hour, minute = divmod(minutes, 60)
        amp = "PM" if hour >= 12 else "AM"
        hour12 = hour if 1 <= hour <= 12 else abs(hour - 12)
        return f"{hour12}:{minute:02d} {amp}"
    return str(val)


def parse_start_minutes(time_str: str) -> int:
    """
    Convert first time in a range ('730', '7:30', '8–9', '08:00 AM') to minutes
    after midnight.  Returns +∞ for unparsable strings (ensures such rows go last).
    """
    if not time_str:
        return math.inf

    s = time_str.lower()
    token = re.split(r'[–-]', s)[0].strip()            # up to first dash
    ampm_in_token = re.search(r'\b(am|pm)\b', token)
    ampm = ampm_in_token.group(1) if ampm_in_token else None

    # Fallback: look later in string for am/pm if not in token
    if ampm is None:
        g = re.search(r'\b(am|pm)\b', s)
        ampm = g.group(1) if g else None

    # Remove am/pm & punctuation for number parsing
    digits = re.sub(r'\b(?:am|pm)\b', '', token)
    digits = digits.replace(':', '').replace(' ', '')

    if not digits.isdigit():
        return math.inf

    if len(digits) <= 2:                    # “8”
        hour, minute = int(digits), 0
    elif len(digits) == 3:                  # “730”
        hour, minute = int(digits[:-2]), int(digits[-2:])
    else:                                   # “0830”, “1030”
        hour, minute = int(digits[:-2]), int(digits[-2:])

    # Apply AM/PM adjustments
    if ampm == 'pm' and hour < 12:
        hour += 12
    elif (ampm == 'am' or ampm is None) and hour == 12:
        hour = 0

    return hour * 60 + minute


def cell_hex(cell) -> str | None:
    """Return '#RRGGBB' from a cell’s fill, ignoring white/black/blank."""
    rgb = getattr(cell.fill.start_color, "rgb", None)
    if rgb and rgb[-6:].lower() not in ("ffffff", "000000"):
        return f"#{rgb[-6:]}"
    return None


def badge_class(setting: str) -> str:
    """Map Excel Setting column to CSS class."""
    return "virtual" if setting.strip().lower().startswith("v") else "in-person"


def next_month_same_day(date: dt.date) -> dt.date:
    """Return the same day next month (or last valid day if month shorter)."""
    year = date.year + (date.month // 12)
    month = (date.month % 12) + 1
    last = calendar.monthrange(year, month)[1]
    return dt.date(year, month, min(date.day, last))


# ────────────────────────── load events from workbook ───────────────────────
def load_events(xlsx: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]                  # first sheet = current academic year

    evts = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        d, t, subj, lect, setting = row[:5]
        if not isinstance(d.value, (dt.datetime, dt.date)):
            continue                           # skip headers/blanks

        date = d.value.date() if isinstance(d.value, dt.datetime) else d.value
        time_str = format_time(t.value)

        evts.append(
            {
                "date":      date,
                "time":      time_str,
                "start_min": parse_start_minutes(time_str),
                "subject":   (subj.value or "").strip(),
                "lecturer":  (lect.value or "").strip(),
                "setting":   (setting.value or "").strip(),
                "row_hex":   cell_hex(subj),
            }
        )
    return evts


# ────────────────────────── page template fragments ─────────────────────────
HEAD = """<!DOCTYPE html>
<html lang="en"><head>
<meta charset="UTF-8"/><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Didactics Calendar – UCI Dermatology</title>
<style>
body{font-family:Arial,Helvetica,sans-serif;margin:0;display:flex}
nav{width:200px;background:#f7f7f7;padding:20px;box-shadow:2px 0 5px rgba(0,0,0,.1)}
nav img{max-width:100%;border-radius:4px;margin-bottom:1rem}
nav a{display:block;margin-bottom:1rem;text-decoration:none;color:#036;font-weight:bold}
main{flex:1;padding:20px}
h1{margin-top:0;font-size:20px}
table{width:100%;border-collapse:collapse;margin-bottom:1rem}
td{border:1px solid #ccc;padding:4px 6px;font-size:.9rem}
.date-bar{background:#333;color:#fff;padding:6px 10px;font-weight:bold;
          display:inline-block;margin-bottom:.3rem;border-radius:4px}
.status{margin-left:8px;padding:2px 6px;border-radius:4px;font-size:.75rem}
.in-person{background:#32cd32}.virtual{background:#1e90ff}
.break-row{font-style:italic;text-align:center;background:#f4f4f4}
</style></head><body>
<nav><img src="https://sidharthchand.github.io/uci-pages/anteaterzotzot.jpeg" alt="Anteater"/>
<a href="index.html">Home</a></nav><main><h1>Didactics Calendar</h1>
<p style="max-width:50em;font-size:.9rem;margin-bottom:40pt;">
Please note the following didactics calendar is intended as a preview of the scheduled didactics for UCI Dermatology and is issued here up to a month in advance.
The final and official didactics schedule is distributed weekly via email at least five days prior to Friday didactics.
Any schedule posted here is subject to change without notification.
</p>"""

FOOT = "</main></body></html>"


# ────────────────────────── render HTML ─────────────────────────────────────
def render(events: list[dict], out_path: Path) -> None:
    out = [HEAD]
    current_date = None

    for idx, ev in enumerate(events):
        if ev["date"] != current_date:                    # = new day header
            current_date = ev["date"]
            pretty = current_date.strftime("%A, %B %-d %Y")

            day_events = [e for e in events if e["date"] == current_date]
            show_badge = any(e["lecturer"] for e in day_events)
            if show_badge:
                badge = badge_class(day_events[0]["setting"])
                badge_txt = "VIRTUAL" if badge == "virtual" else "IN PERSON"
                out.append(f'<span class="date-bar">{pretty}<span class="status {badge}">{badge_txt}</span></span>')
            else:
                out.append(f'<span class="date-bar">{pretty}</span>')

            out.append("<table>")

        subj_style = f' style="background:{ev["row_hex"]};"' if ev["row_hex"] else ""
        if not ev["lecturer"]:                            # break/holiday row
            out.append(f'<tr><td class="break-row" colspan="3"{subj_style}>{html.escape(ev["subject"])}</td></tr>')
        else:                                             # normal lecture row
            out.append(
                f'<tr><td>{html.escape(ev["time"])}</td>'
                f'<td{subj_style}>{html.escape(ev["subject"])}</td>'
                f'<td>{html.escape(ev["lecturer"])}</td></tr>'
            )

        # close table when next event is a different date
        nxt_date = events[idx + 1]["date"] if idx + 1 < len(events) else None
        if nxt_date != current_date:
            out.append("</table>")

    out.append(FOOT)
    out_path.write_text("\n".join(out), encoding="utf-8")
    print(f"✓ Wrote {out_path}")


# ────────────────────────── main ────────────────────────────────────────────
def main() -> None:
    ap = argparse.ArgumentParser(description="Excel → styled didactics calendar")
    ap.add_argument("excel", help="Path to didactics workbook (.xlsx)")
    ap.add_argument("output_html", nargs="?", default="didactics_calendar.html")
    args = ap.parse_args()

    events = load_events(Path(args.excel))

    today  = dt.date.today()
    window_start = today.replace(day=1)           # always start 1st of this month
    window_end   = next_month_same_day(today)     # inclusive same-day next month

    events = [e for e in events if window_start <= e["date"] <= window_end]
    if not events:
        sys.exit("No events found in this 1-month window.")

    # Sort by date, then chronological start time
    events.sort(key=lambda e: (e["date"], e["start_min"]))

    render(events, Path(args.output_html))


if __name__ == "__main__":
    main()
